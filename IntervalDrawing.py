
import requests
from requests.exceptions import MissingSchema
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.common.exceptions import NoSuchElementException
from selenium.common.exceptions import ElementClickInterceptedException
from selenium.common.exceptions import InvalidSelectorException
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome import service as cs
from selenium.webdriver.common.keys import Keys
import time
import cv2
import tempfile
import os
from selenium.webdriver.support import expected_conditions as EC
import openpyxl
import random
import threading 
import win32gui
import numpy as np
import tkinter as tk
import csv
import queue



INTERVAL = 1
RETRY_NUM = 300
TIMER_INTERVAL = 10
TIMER_WIDTH = 100
TIMER_HEIGHT = 50
URL = "https://www.google.co.jp/imghp?hl=ja"
driver_path = "./chromedriver"

X_in_show_info = 1
Y_in_show_info = 2
height_in_show_info = 3

FINISH_COMMAND = False
CONFIRM = False
img_urls_previous = []
miss_num = 0
missing_count = 0
word_column = 0


def search_xlsx():
    for file in os.listdir():
        base, ext = os.path.splitext(file)
        if ext == ".xlsx":
            print(file)
            return file

def imread_web(url):#urlから画像を一時保存
    res = requests.get(url)
    img = None
    fp = tempfile.NamedTemporaryFile(dir='./', delete=False)
    time.sleep(1)
    #print(fp.name)
    fp.write(res.content)
    fp.close()
    img = cv2.imread(fp.name)
    os.remove(fp.name)
    return img

def click(driver,elem):#スクロールしてクリック
    driver.execute_script("arguments[0].scrollIntoView(false);", elem)
    elem.click()

def headless_option():
    options = Options()
    options.add_argument('--headless')
    options.add_argument('--no-sandbox')
    options.add_argument('--disable-dev-shm-usage')
    # options.add_argument('--start-maximized')
    options.add_argument('--start-fullscreen')
    options.add_argument('--disable-plugins')
    options.add_argument('--disable-gpu')
    options.add_argument('--disable-extensions')
    options.add_argument('--ignore-certificate-errors')
    options.add_argument('--ignore-ssl-errors')
    return options

def get_urls(driver):#chromeから画像のurlをexcelに保存 #countの他もう一つ、excelに保存するときに飛びがでないようなindexを入れる
    count = 0#漏れがあったら飛ばせるようにするindex
    index = 0#excel保存用のindex
    while(True):
        xpath = "//img"
        tmb_imgs = driver.find_elements(by=By.XPATH, value=xpath)
        tmb_alts = [tmb.get_attribute('alt') for tmb in tmb_imgs]
        for tmb_img,tmb_alt in zip(tmb_imgs, tmb_alts):
            try:
                count = count+1
                if count == 1:
                    continue
                if tmb_alt == '':
                    count = count -1
                    continue

                click(driver,tmb_img)
                img_elem = driver.find_element(by=By.ID, value = 'islsp')
                tmb_url = tmb_img.get_attribute('src')  # サムネイル画像のsrc属性値
                img_img = img_elem.find_element(by = By.CSS_SELECTOR, value=f'img[alt=\'{tmb_alt}\']')
                
                for i in range(RETRY_NUM):
                    img_url = img_img.get_attribute('src')
                    if img_url == tmb_url:  # src属性値が遷移するまでリトライ
                        time.sleep(0.1)
                        img_url = ''
                    else:
                        continue

                if img_url.startswith('http')==True and img_url.lower().find('gif')==-1:
                    #print(img_url)
                    index = index+1
                    print(index)
                    sheet.cell(index+1,word_column).value = img_url
                    if index>num-1:
                        break
                else:
                    continue
            #except InvalidSelectorException:
            except Exception as e:
                print(e)
                continue
            
        if index>num-1:
            break
        driver.execute_script('window.scrollTo(0, document.body.scrollHeight);')
        time.sleep(1)
    
    wb.save(exfile)
    

def randint_from_a_to_b_in_range_k(a, b, k):
        turn_list = []
        while len(turn_list) < k:
            n = random.randint(a, b)
            if not n in turn_list:
                turn_list.append(n)
        return turn_list
        

def show_img_from_url_at_random(img_urls,n):
    global miss_num
    global missing_count
    img_index = 0
    if n < 0:
        return
    while(1):
        try:
            open_img = imread_web(img_urls[img_index])
            
            height = open_img.shape[0]
            width = open_img.shape[1]
            resize_height = show_info.cell(2,height_in_show_info).value
            resize_weight = resize_height/height
            resize_width = width*resize_weight
            open_img_resize = cv2.resize(open_img,(int(resize_width),int(resize_height)))
            
            move_window_x = show_info.cell(2,X_in_show_info).value
            move_window_y = show_info.cell(2,Y_in_show_info).value
            
            real_index = img_index+1
            cv2.putText(
                open_img_resize,
                text = f'{real_index}',
                org=(2, 27),
                fontFace=cv2.FONT_HERSHEY_TRIPLEX | cv2.FONT_ITALIC,
                fontScale=1.0,
                color=(0,0,0),#明度変えて馴染ませたかったけど、処理時間が問題
                thickness=2,
                lineType=cv2.LINE_AA
            )
            cv2.imshow("img",open_img_resize)
            cv2.moveWindow("img", move_window_x,move_window_y)
            img_find = win32gui.FindWindow(None, 'img')
            if img_find:
                recta = win32gui.GetWindowRect(img_find)
                print(recta)
            print(f"URL({real_index}): ",img_urls[img_index])
            t = False
            print("0:00",end="")
            for i in range(int(interv*60/TIMER_INTERVAL)):
                

                k = cv2.waitKey(TIMER_INTERVAL*1000) 
                if k != -1:
                    if k == ord('q'):
                        t=True
                        break
                    else:
                        break
                if i%6==5:#?:00の形にする
                    print(f"\r{divmod(i+1,6)[0]}:{divmod(i+1,6)[1]*TIMER_INTERVAL}0", end="")
                else:
                    print(f"\r{divmod(i+1,6)[0]}:{divmod(i+1,6)[1]*TIMER_INTERVAL}", end="")

            print("")
            if t==True:
                break
            img_index = img_index+1
            if img_index > n -1 + miss_num:
                break

        except IndexError:
            break
        except AttributeError:
            time.sleep(0.1)
            continue
        except MissingSchema:
            img_index = img_index + 1
            miss_num = miss_num+1
            missing_count = missing_count+1
            if missing_count > 5:
                break
seconds =0
minutes =0

def timer():
    root = tk.Tk()

    # Set window title 
    root.title("Stopwatch") 

    # Set window size 
    root.geometry("220x70") 

    # Define global variables

    # Function to increase stopwatch time by one second
    def increment_time(): 
        global seconds 
        global minutes 
        # Increase seconds 
        seconds += 1
        # When seconds reach 60, increase minutes 
        if seconds == 60: 
            minutes += 1
            seconds = 0
        # Display updated time 
        stopwatch.config(text = "{}:{}".format(minutes, seconds))
        # Call the increment_time function after 1 second
        stopwatch.after(1000, increment_time) 

    # Create a label to display the stopwatch 
    stopwatch = tk.Label(root, text = "0:0", font = ("Arial Bold", 20)) 
    stopwatch.pack()

    # Call the increment_time function 
    stopwatch.after(1000, increment_time) 

    # Start the main event loop
    root.mainloop()


def show_img_from_url(img_urls,n):
    print("You can finish this pragram if you input \"q\"")
    print("And you can skip the picture if you input ANY KEYS except \"q\"")
    print("ready...3")
    time.sleep(2)
    print("ready...2")
    time.sleep(2)
    print("ready...1")
    time.sleep(1)
    show_img_from_url_at_random(img_urls,n)
    cv2.destroyWindow("img")
    print("\nEnd")

def word_place(word):
    print(word)
    global word_column
    label = sheet[1]#一行目（ラベル）
    word_exist = False
    for i in range(len(label)):
        if word == sheet.cell(1,i+1).value:
            word_column = i+1
            word_exist = True
            break
    
    if word_exist == False:
        word_column = len(label)+1
        new_word_cell = sheet.cell(1,word_column)
        new_word_cell.value = word

def collect_img_urls_previous(num,word):
    img_urls = []
    print("now collectiong...")
    word_place(word)
    print("word_column",word_column)
    word_row=0
    for i in range(len(list(sheet.columns)[word_column-1])):#listにしたからindexは0から
        if sheet.cell(i+2,word_column).value == None:#sheetはindexが1から
            word_row = i+1
            break
    print("max_num",word_row-1)
    turn_list = randint_from_a_to_b_in_range_k(2,word_row,num)
    print("pick_up_index",turn_list)
    for i in turn_list:
        img_urls.append(sheet.cell(i,word_column).value) 
    return img_urls  


def getSizeCallback(event):
    show_info.cell(2,X_in_show_info).value = event.x
    show_info.cell(2,Y_in_show_info).value = event.y
    show_info.cell(2,height_in_show_info).value = event.height
    print(f"x:{event.x}, y:{event.y}, height:{event.height}")
    
def get_img_coordinate():   
    win = tk.Tk()
    x = show_info.cell(2,X_in_show_info).value
    y = show_info.cell(2,Y_in_show_info).value
    width = show_info.cell(2,height_in_show_info).value
    height = width
    win.geometry(f"{width}x{height}+{x}+{y}") 
    win.bind("<Configure>",getSizeCallback)
    win.mainloop()
    wb.save(exfile)


if __name__ == "__main__":
    
    try:
        exfile = search_xlsx()
        print("Excel file:", end = "")
        print(exfile)
        wb = openpyxl.load_workbook(exfile)
        try:
            sheet  = wb["URLdatabase"]
        except:
            wb.create_sheet(index = 0, title = "URLdatabase")
            sheet  = wb["URLdatabase"]
        try:
            show_info = wb["Information"]
        except:
            wb.create_sheet(index = 0, title = "Information")
            show_info = wb["Information"]
        while(1):
            print("List up:")
            for i in range(len(sheet[1])):
                print(f"{sheet.cell(1,i+1).value} / ", end="")
            print("")
            print("Input \"g\", \"s\" ,\"a\",or \"f\"\ng:get URLs  s:show images a:adjust size & posision of window f:finish pragram")
            get_or_show = input()

            if get_or_show == "g":
                print("Input the Search Words. Separete with \"・\"")
                words = input()
                word_list = words.split("・")
                while(1):
                    try:
                        print("Input each number of items")
                        num = input()
                        num = int(num)
                        break
                    except ValueError:
                        print("Error: Input Integer")
                print("Are you OK? if so, press ENTER. If you'll redo the setup, input any other words")
                ok = input()
                if ok=="":
                    pass
                else:
                    continue
                print("Setup is complete")
                
                chrome_service = cs.Service(executable_path="./chromedriver")
                #driver = webdriver.Chrome(service = chrome_service, options=headless_option())
                driver = webdriver.Chrome(service = chrome_service)
                
                for word in word_list:
                    driver.get(URL)
                    time.sleep(INTERVAL)
                    driver.find_element(by =By.NAME,value = 'q').send_keys(word,Keys.ENTER)
                    time.sleep(INTERVAL)
                    word_place(word)
                    get_urls(driver)
                driver.close()
                print("driver_closed")
                break
            
            elif get_or_show == "s":
                label_list = []
                for i in range(len(sheet[1])):
                    label_list.append(sheet.cell(1,i+1).value)
                while(1):
                    print("Input the Search Words")
                    word = input()
                    if not word in label_list:
                        print("Error: Choose from List above")
                    else:
                        break
                while(1):
                    try:
                        print("Input the Number of Times")
                        num = input()
                        num = int(num)
                        break
                    except ValueError:
                        print("Error: Input Integer")
                while(1):
                    try:
                        print("Input How Long the Interval is [min]")
                        interv = input()
                        interv = float(interv)
                        break
                    except ValueError:
                        print("Error: Input Float")
                print("Are you OK? if so, press ENTER. if you'll redo the setup, input any other words")
                ok = input()
                if ok=="":
                    pass
                else:
                    continue
                print("Setup is complete")
                img_urls_previous = collect_img_urls_previous(num,word)
                
                show_img_from_url(img_urls_previous,num)
                break
            elif get_or_show == "a":
                print("Adjust the place of \"example\" for S_mode")
                get_img_coordinate()
            elif get_or_show == "f":
                break
            else:
                print("Invalid Input")
                time.sleep(1)
    except Exception as e:
        print("Error: ", end = "")
        print(e)


    
    
